import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

try:
    wb = openpyxl.load_workbook('EduMap_Documentation.xlsx')
    sheet = wb['5. Database Design']
    
    headers = [cell.value for cell in sheet[1]]
    
    with open('sheet5_full_details.txt', 'w', encoding='utf-8') as f:
        f.write("Headers: " + " | ".join([h for h in headers if h is not None]) + "\n\n")
        
        for r in range(2, sheet.max_row + 1):
            f.write(f"--- Row {r-1} ---\n")
            for col_idx, h in enumerate(headers):
                cell_val = sheet.cell(row=r, column=col_idx + 1).value
                if cell_val is not None:
                    f.write(f"{h}: {cell_val}\n")
            f.write("\n")
            
    print("Dumped Sheet 5 details successfully!")
except Exception as e:
    print("Error:", str(e))
