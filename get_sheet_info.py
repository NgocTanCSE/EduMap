import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

try:
    wb = openpyxl.load_workbook('EduMap_Documentation.xlsx')
    for index, name in enumerate(wb.sheetnames):
        sheet = wb[name]
        print(f"Sheet {index}: {name} | Rows: {sheet.max_row} | Cols: {sheet.max_column}")
except Exception as e:
    print("Error:", str(e))
