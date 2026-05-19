import openpyxl
import os

os.makedirs('scratch', exist_ok=True)
wb = openpyxl.load_workbook('EduMap_Documentation.xlsx')

for idx, name in enumerate(wb.sheetnames, 1):
    sheet = wb[name]
    filename = f"scratch/sheet_{idx}_{name.replace('/', '_').replace(' ', '_')}.txt"
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(f"--- SHEET: {name} ---\n\n")
        # Dump first 100 rows to avoid extremely large files, usually documentation is small
        for r in range(1, min(sheet.max_row + 1, 150)):
            row_vals = [cell.value for cell in sheet[r]]
            if any(val is not None for val in row_vals):
                row_str = " | ".join([str(val).strip().replace('\n', ' ') if val is not None else "" for val in row_vals])
                f.write(f"Row {r}: {row_str}\n")
    print(f"Dumped sheet {idx}")
