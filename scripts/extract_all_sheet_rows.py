import openpyxl
import os

def analyze():
    wb = openpyxl.load_workbook('EduMap_Documentation.xlsx')
    output_path = 'scripts/all_sheets_data.txt'
    with open(output_path, 'w', encoding='utf-8') as f:
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            f.write(f"\n=========================================\n")
            f.write(f"SHEET: {sheet_name} (Rows: {sheet.max_row}, Cols: {sheet.max_column})\n")
            f.write(f"=========================================\n")
            for r in range(1, sheet.max_row + 1):
                row_vals = [cell.value for cell in sheet[r]]
                # Filter out completely empty rows
                if any(val is not None for val in row_vals):
                    row_str = " | ".join([str(val) if val is not None else "" for val in row_vals])
                    f.write(f"Row {r}: {row_str}\n")
    print(f"Analysis completed! All data written to {output_path}")

if __name__ == "__main__":
    analyze()
