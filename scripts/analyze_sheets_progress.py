import openpyxl
import os

def analyze():
    wb = openpyxl.load_workbook('EduMap_Documentation.xlsx')
    print("Sheets in workbook:")
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} (Rows: {sheet.max_row}, Cols: {sheet.max_column}) ---")
        # Print top 5 rows
        for r in range(1, min(6, sheet.max_row + 1)):
            row_vals = [cell.value for cell in sheet[r]]
            print(f"Row {r}: {row_vals[:8]}")

if __name__ == "__main__":
    analyze()
