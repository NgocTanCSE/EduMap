import openpyxl
import sys

try:
    wb = openpyxl.load_workbook('EduMap_Documentation.xlsx')
    sheet_name = wb.sheetnames[2] # 3. Danh Sách Chức Năng
    sheet = wb[sheet_name]
    
    with open('sheet3_content.txt', 'w', encoding='utf-8') as f:
        f.write(f"--- Sheet: {sheet_name} ---\n\n")
        for r in range(1, sheet.max_row + 1):
            row_vals = [cell.value for cell in sheet[r]]
            if any(row_vals is not None for row in sheet[r]):
                formatted_row = " | ".join([str(val) if val is not None else "" for val in row_vals])
                f.write(formatted_row + "\n")
    print("Successfully written Sheet 3 content to sheet3_content.txt")
except Exception as e:
    print("Error:", str(e))
