import openpyxl
import sys

try:
    wb = openpyxl.load_workbook('EduMap_Documentation.xlsx')
    
    # Let's dump Sheet 1 (2. Danh Sách Module)
    s1 = wb['2. Danh Sách Module']
    with open('sheet1_module_dump.txt', 'w', encoding='utf-8') as f:
        for r in range(1, s1.max_row + 1):
            row_vals = [cell.value for cell in s1[r]]
            f.write(" | ".join([str(val) if val is not None else "" for val in row_vals]) + "\n")
            
    # Let's dump Sheet 2 (3. Chi Tiết Chức Năng)
    s2 = wb['3. Chi Tiết Chức Năng']
    with open('sheet2_function_dump.txt', 'w', encoding='utf-8') as f:
        for r in range(1, s2.max_row + 1):
            row_vals = [cell.value for cell in s2[r]]
            f.write(" | ".join([str(val) if val is not None else "" for val in row_vals]) + "\n")
            
    print("Dumps created successfully!")
except Exception as e:
    print("Error:", str(e))
